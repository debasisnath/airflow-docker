#! /usr/bin/env python3

import os
import sys
from datetime import datetime, timedelta

import time
from functools import wraps

from string import Template

import pandas as pd
from pyhive import hive
from impala.dbapi import connect
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule

import asyncio

from DQI_config_vw import *


def timeit(func):
    @wraps(func)
    def timeit_wrapper(*args, **kwargs):
        start_time = time.perf_counter()
        if asyncio.iscoroutinefunction(func):
            async def async_wrapper():
                result = await func(*args, **kwargs)
                end_time = time.perf_counter()
                total_time = end_time - start_time
                print(f'Function {func.__name__}{args} {kwargs} Took {total_time:.4f} seconds')
                return result
            return async_wrapper()
        else:
            result = func(*args, **kwargs)
            end_time = time.perf_counter()
            total_time = end_time - start_time
            print(f'Function {func.__name__}{args} {kwargs} Took {total_time:.4f} seconds')
            return result
    return timeit_wrapper


class ReportGenerator:
    def __init__(self, sql_file_path, xlsx_path, sql_context, create_new=True):
        # Store stateful configuration and context variables.
        self.sql_file_path = sql_file_path
        self.xlsx_path = xlsx_path
        self.sql_context = sql_context

        # Initialize connections as None; they'll be set later.
        self.hive_conn = None
        self.impala_conn = None

        # Update SQL context with current date if needed.
        self.full_date, self.current_datetime = self.get_current_dates()
        self.sql_context.setdefault("full_dte", self.full_date)

        # make connections ready
        self.setup_connections()

        # ✨✨✨
        if create_new:
            loop = asyncio.get_event_loop()
            loop.run_until_complete(self.create_new_tables())

        # Initialize connection-dependent variables
        self.initialize_vars()

    # ✨✨✨
    async def create_new_tables(self):
        await self.drop_views()
        await self.create_table_1()
        await self.create_table_2()

    def initialize_vars(self):
        # set vars(connection dependent)
        # name
        self.total_records_value = self.get_total_records_value()
        self.missing_record_customer_name = self.get_missing_record_customer_name()
        self.records_passed_customer_name_validity = self.get_records_passed_customer_name_validity()
        self.records_failed_customer_name_validity = self.get_records_failed_customer_name_validity()

        # dob/age
        self.missing_record_age_dob = self.get_missing_record_age_dob()
        self.records_passed_age_validity = self.get_records_passed_age_validity()
        self.records_failed_age_validity = self.get_records_failed_age_validity()

        # identifier
        self.missing_records_identifier = self.get_missing_records_identifier()
        # self.records_passed_identifier_validity = self.get_records_passed_identifier_validity()
        self.records_failed_identifier_validity = self.get_records_failed_identifier_validity()

        # phone
        self.missing_records_phone_number = self.get_missing_records_phone_number()
        self.records_failed_phone_number_validity = self.get_records_failed_phone_number_validity()

        # address
        self.missing_record_address = self.get_missing_record_address()
        self.records_failed_address_validity = self.get_records_failed_address_validity()

        # total_income
        self.missing_record_total_income = self.get_missing_record_total_income()
        self.records_failed_total_income_validity = self.get_records_failed_total_income_validity()

        # key_person
        self.missing_record_key_person = self.get_missing_record_key_person()
        self.records_failed_key_person_validity = self.get_records_failed_key_person_validity()

    @timeit # ✨✨✨
    async def drop_views(self):
        print('@drop_views 1')
        sql_query1 = self.get_sql_from_file('drop_vw_table.sql')
        self.execute_query(self.impala_conn, sql_query1)
        print(f"Executed query: {sql_query1}")

        print('@drop_views 2')
        sql_query2 = self.get_sql_from_file('drop_vw_table_curated.sql')
        self.execute_query(self.impala_conn, sql_query2)
        print(f"Executed query: {sql_query2}")

    @timeit # ✨✨✨
    async def create_table_1(self):
        print('@create_new 1')
        sql_query1 = self.get_sql_from_file('create_vw_table.sql')
        self.execute_query(self.impala_conn, sql_query1)
        print(f"Executed query1: ")

    @timeit # ✨✨✨
    async def create_table_2(self):
        print('@create_new 2')
        sql_query2 = self.get_sql_from_file('create_vw_table_curated.sql')
        self.execute_query(self.impala_conn, sql_query2)
        print(f"Executed query2: ")

    def setup_connections(self):
        """
        Establishes and stores the Hive and Impala connections on the instance.
        """
        try:
            self.hive_conn = hive.Connection(
                host=HIVE_HOST,
                port=HIVE_PORT,
                database=HIVE_DATABASE,
                username=HIVE_USERNAME
            )
            hive.connect(HIVE_HOST, configuration=HIVE_CONFIG)
        except Exception as e:
            print("Hive connection is not established due to exception ::", repr(e))

        try:
            self.impala_conn = connect(host=IMPALA_HOST, port=IMPALA_PORT)
        except Exception as e:
            print("Impala connection is not established due to exception ::", repr(e))

    def get_current_dates(self):
        """
        Parses command-line arguments to get a full date.
        If not provided, uses the date from two days ago.
        Returns:
            tuple: (full_date [yyyy-mm-dd], current_datetime_str [yyyymmdd-HHMMSS])
        """
        try:
            full_date = sys.argv[1]
            current_datetime = f"{full_date}-{datetime.now().strftime('%H%M%S')}"
        except Exception:
            dt_val = datetime.now() - timedelta(days=2)
            full_date = dt_val.strftime('%Y-%m-%d')
            current_datetime = dt_val.strftime('%Y%m%d-%H%M%S')
        print('full_dte ->', full_date)
        return full_date, current_datetime

    def get_sql_from_file(self, file_name):
        """
        Reads SQL from a file located in the stateful sql_file_path and substitutes
        placeholders using the instance's sql_context dictionary.
        """
        file_path = os.path.join(self.sql_file_path, file_name)
        with open(file_path, 'r') as file:
            sql_template = file.read()
        template = Template(sql_template)
        return template.safe_substitute(self.sql_context)

    def get_query_result(self, sql):
        """
        Executes the provided SQL query using the instance's Impala connection
        and returns the result as a pandas DataFrame.
        """
        return pd.read_sql(sql, self.impala_conn)

    @timeit
    def get_total_records_value(self):
        """
        Fetches the total records count by reading the specific SQL and executing it.
        Assumes that the query returns a single value.
        """
        sql_query = self.get_sql_from_file('total-records-count.sql')
        print("Executed SQL Query:\n", sql_query)
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        total_records_value = df.iloc[0, 0] if not df.empty else 0
        return total_records_value

    @timeit
    def get_missing_record_customer_name(self):
        print('@missing_record_customer_name')
        sql_query = self.get_sql_from_file('missing-record-customer-name.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_records_passed_customer_name_validity(self):
        print('@records_passed_customer_name_validity')
        sql_query = self.get_sql_from_file('records-passed-customer-name-validity.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_records_failed_customer_name_validity(self):
        print('@records_failed_customer_name_validity')
        sql_query = self.get_sql_from_file('records-failed-customer-name-validity.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_missing_record_age_dob(self):
        print('@get_missing_record_age_dob')
        sql_query = self.get_sql_from_file('missing-record-age-dob.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_records_passed_age_validity(self):
        print('@get_records_passed_age_validity')
        sql_query = self.get_sql_from_file('records-passed-age-validity.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_records_failed_age_validity(self):
        print('@get_records_failed_age_validity')
        sql_query = self.get_sql_from_file('records-failed-age-validity.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_missing_records_identifier(self):
        print('@get_missing_records_identifier')
        sql_query = self.get_sql_from_file('missing-record-identifier.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_records_passed_identifier_validity(self):
        print('@get_records_passed_identifier_validity')
        sql_query = self.get_sql_from_file('records-passed-identifier-validity.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_records_failed_identifier_validity(self):
        print('@get_records_failed_identifier_validity')
        sql_query = self.get_sql_from_file('records-failed-identifier-validity.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_missing_records_phone_number(self):
        print('@get_missing_records_phone_number')
        sql_query = self.get_sql_from_file('missing-records-phone-number.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_records_failed_phone_number_validity(self):
        print('@get_records_failed_phone_number_validity')
        sql_query = self.get_sql_from_file('records-failed-phone-number-validity.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_missing_record_address(self):
        print('@get_missing_record_address')
        sql_query = self.get_sql_from_file('missing-record-address.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_records_failed_address_validity(self):
        print('@get_records_failed_address_validity')
        sql_query = self.get_sql_from_file('records-failed-address-validity.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_missing_record_total_income(self):
        print('@get_missing_record_total_income')
        sql_query = self.get_sql_from_file('missing-record-total-income.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_records_failed_total_income_validity(self):
        print('@get_records_failed_total_income_validity')
        sql_query = self.get_sql_from_file('records-failed-total-income-validity.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_missing_record_key_person(self):
        print('@get_missing_record_key_person')
        sql_query = self.get_sql_from_file('missing-record-key-person.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    @timeit
    def get_records_failed_key_person_validity(self):
        print('@get_records_failed_key_person_validity')
        sql_query = self.get_sql_from_file('records-failed-key-person-validity.sql')
        df = self.get_query_result(sql_query)
        print("Query result:\n", df.head())
        _value = df.iloc[0, 0] if not df.empty else 0
        return _value

    def _build_row(self, 
            sub_category, 
            category, 
            dq_steward, 
            database_name, 
            table_names,
            param_fields, 
            field_desc,
            month,
            total_records,
            missing_records,
            total_qualified, 
            passed, 
            failed, 
            pass_percentage, 
            current_score,
            expected_score, 
            threshold,
            dimension,
            rule_desc,
            pass_fail="Pass",
            failed_location="NA"
        ):
        """
        Helper method that returns a row (list) with the fixed column order.
        The first column (Sno) is left as None to be auto-generated later.
        Columns:
         0: Sno, 1: SUB-CATEGORY, 2: CATEGORY, 3: DQ Steward, 4: Data base name,
         5: Table Name(s), 6: Parameters/Field Name(s), 7: Field Description, 8: Month,
         9: Total Records, 10: Missing Records, 11: Total Records Qualified,
         12: Records passed, 13: Records Failed, 14: Pass %, 15: Current Score,
         16: Expected Score, 17: Threshold, 18: Data Quality Dimension,
         19: Data Quality Rule description, 20: Pass/Fail, 21: Failed record location.
        """
        return [
            None,
            sub_category,
            category,
            dq_steward,
            database_name,
            table_names,
            param_fields,
            field_desc,
            month,
            total_records,
            missing_records,
            total_qualified,
            passed,
            failed,
            pass_percentage,
            current_score,
            expected_score,
            threshold,
            dimension,
            rule_desc,
            pass_fail,
            failed_location,
        ]

    def _add_row_wise_conditional_formatting(self, ws, headers):
        """
        Adds a conditional formatting rule that highlights the entire row 
        if the 'Data Quality Dimension' column's value is 'Validity'.
        This method follows the Single Responsibility Principle.
        """
        if "Data Quality Dimension" not in headers:
            # Optionally log or print a warning here.
            return
        else:
            # Identify the column for "Data Quality Dimension".
            desc_index = headers.index("Data Quality Dimension") + 1  # 1-indexed.
            desc_col_letter = get_column_letter(desc_index)
            last_col_letter = get_column_letter(len(headers))
            
            # Define the range for all data rows.
            data_range = f"A2:{last_col_letter}{ws.max_row}"
            
            # Use the INDIRECT function to dynamically refer to the cell in the current row.
            # This formula will be interpreted relative to each row in the `data_range`.
            rule_formula = f'=INDIRECT("${desc_col_letter}" & ROW())="Validity"'

            # Define a new fill style for the entire row.
            # If the color seems too subtle, try a more contrasting color (e.g., "FFFF00" for bright yellow).
            # row_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            row_fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
            differential_style = DifferentialStyle(fill=row_fill)
            
            rule = Rule(type="expression", dxf=differential_style)
            rule.formula = [rule_formula]
            
        ws.conditional_formatting.add(data_range, rule)
    
    def get_rows(self):
        """
        Assembles and returns rows (list of lists) for the Excel report.
        The rows include dummy data along with query results and the current user.
        """
        current_user = os.getlogin()
        print("CURRENT USER ::", current_user)
        # total_records_value = self.get_total_records_value()

        # report_month = REPORT_MONTH
        report_month = POST_TXT
        
        # Note: The first element `Sno`, the `Pass %` ,,, columns will be overridden.
        # ✅ Name (Validity)
        row01 = self._build_row(
            sub_category  = "Demographic Parameters",
            category = "Name",
            dq_steward = current_user,
            database_name = "BBLEDL",
            table_names = "tb_cic_cif_snp_edl",
            param_fields = "nam",
            field_desc = "Customer Name",
            month = report_month,
            total_records = self.total_records_value,
            missing_records = self.missing_record_customer_name,
            total_qualified = self.total_records_value,
            passed = self.records_passed_customer_name_validity,
            failed = self.records_failed_customer_name_validity,
            pass_percentage = None,                                 # `Pass %` will use formula.
            current_score = None,                                   # `Current Score` will use formula
            expected_score = EXPECTED_SCORE_CUSTOMER_NAME,
            threshold = None,                                       # `Threshold` defined in config
            dimension = "Validity",
            rule_desc = ("Availability which satisfies all conditions: a) Minimum 2 tokens, "
             "b) 1 token with minimum 2 letters of the alphabet, c) No numerals present"),
        )

        # ✅ Name (Completeness)
        row02 = self._build_row(
            sub_category  = "Demographic Parameters",
            category = "Name",
            dq_steward = current_user,
            database_name = "BBLEDL",
            table_names = "tb_cic_cif_snp_edl",
            param_fields = "nam",
            field_desc = "Customer Name",
            month = report_month,
            total_records = self.total_records_value,
            missing_records = self.missing_record_customer_name,
            total_qualified = self.total_records_value,
            passed = (self.total_records_value - self.missing_record_customer_name),
            failed = self.missing_record_customer_name,
            pass_percentage = None,
            current_score = None,
            expected_score = EXPECTED_SCORE_CUSTOMER_NAME,
            threshold = None,
            dimension = "Completeness",
            rule_desc = "Field cannot be Blank"
        )
        
        # ✅ DOB/AGE (Validity)
        row03 = self._build_row(
            sub_category = "Demographic Parameters",
            category = "DOB/AGE",
            dq_steward = current_user,
            database_name = "BBLEDL",
            table_names = "tb_cic_cif_snp_edl",
            param_fields = "dob , age",
            field_desc = "Customer AGE, date of birth",
            month = report_month,
            total_records = self.total_records_value,
            missing_records =  self.missing_record_age_dob,
            total_qualified = self.total_records_value - self.missing_record_age_dob,
            passed = self.records_passed_age_validity,
            failed = self.records_failed_age_validity,
            pass_percentage = None,
            current_score = None,
            expected_score = EXPECTED_SCORE_AGE_DOB,
            threshold = None,
            dimension = "Validity",
            rule_desc = ("Availability which satisfies all conditions: " 
             "a) If DOB is shared then date format should be DDMMYYYY "
             "b) If Age is provided then it needs to be numeric value except Zero")
        )

        # ✅ Row 4: DOB/AGE (Completeness)
        row04 = self._build_row(
            sub_category="Demographic Parameters",
            category="DOB/AGE",
            dq_steward=current_user,
            database_name="BBLEDL",
            table_names="tb_cic_cif_snp_edl",
            param_fields="dob, age",
            field_desc="Customer AGE, date of birth",
            month=report_month,
            total_records=self.total_records_value,
            missing_records=self.missing_record_age_dob,
            total_qualified=(self.total_records_value - self.missing_record_age_dob),
            passed=(self.total_records_value - self.missing_record_age_dob),
            failed=self.missing_record_age_dob,
            pass_percentage=None,
            current_score= None,
            expected_score = EXPECTED_SCORE_AGE_DOB,
            threshold=None,
            dimension="Completeness",
            rule_desc="Field cannot be Blank"
        )

        # ✅ Row 5: Identifier: Voter ID (VID)/PAN/CKYC (Validity)
        row05 = self._build_row(
            sub_category="Demographic Parameters",
            category="Identifier: Voter ID (VID)/PAN/CKYC",
            dq_steward=current_user,
            database_name="BBLEDL",
            table_names="tb_cic_cif_snp_edl",
            param_fields="voter_id,pan_number,ckyc_number",
            field_desc="Customer's voter_id,pan_number,ckyc_number respectively",
            month=report_month,
            total_records=self.total_records_value,
            missing_records=self.missing_records_identifier,
            total_qualified=(self.total_records_value - self.missing_records_identifier),
            passed=(self.total_records_value -self.records_failed_identifier_validity),              # sample value
            failed=self.records_failed_identifier_validity,
            pass_percentage=None,
            current_score=None,
            expected_score=8,
            threshold=None,
            dimension="Validity",
            rule_desc=("Availability of ANY ONE identifier which satisfies ALL respective conditions: "
                       "1.PAN: a) Should be 10 in length b) First 5 and last character should be alphabets "
                       "c) The 4th character has to be either P or H d) The 6th to 9th character should be numerals "
                       "2.Voter ID : a) Should be between 8 - 16 in length after CIC removes the special characters "
                       "b) First 2 / 3 digits should be alphabets "
                       "3.CKYC: a) Should be 14 in length b) Should be all numeric")
        )

        # ✅ Row 6: Identifier: Voter ID (VID)/PAN/CKYC (Completeness)
        row06 = self._build_row(
            sub_category="Demographic Parameters",
            category="Identifier: Voter ID (VID)/PAN/CKYC",
            dq_steward=current_user,
            database_name="BBLEDL",
            table_names="tb_income_details",
            param_fields="voter_id,pan_number,ckyc_number",
            field_desc="Customer's voter_id,pan_number,ckyc_number respectively",
            month=report_month,
            total_records=self.total_records_value,
            missing_records=self.missing_records_identifier,
            total_qualified=self.total_records_value,
            passed=(self.total_records_value-self.missing_records_identifier),
            failed=self.missing_records_identifier,
            pass_percentage=None,
            current_score=None,
            expected_score=8,
            threshold=None,
            dimension="Completeness",
            rule_desc="Field cannot be Blank"
        )

        # ✅ Row 7: Phone Number (Validity)
        row07 = self._build_row(
            sub_category="Demographic Parameters",
            category="Phone Number",
            dq_steward=current_user,
            database_name="BBLEDL",
            table_names="tb_income_details",
            param_fields="zmph",
            field_desc="Mobile Phone",
            month=report_month,
            total_records=self.total_records_value,
            missing_records=self.missing_records_phone_number,
            total_qualified=(self.total_records_value-self.missing_records_phone_number),
            passed=(self.total_records_value - self.records_failed_phone_number_validity),
            failed=self.records_failed_phone_number_validity,
            pass_percentage=None,
            current_score=None,
            expected_score=7,
            threshold=None,
            dimension="Validity",
            rule_desc=("Availability of mobile number which satisfies all conditions:" 
                       "a) Should be minimum 10 numerals in length "
                       "b) The first digit needs to start with 6,7,8 and 9 "
                       "c) Mobile number straight descending/ascending sequence "
                       "(E.g., 8765432/2345678) or same digits (E.g., 2222222) not allowed")
            )

        # ✅ Row 8: Phone Number (Completeness)
        row08 = self._build_row(
            sub_category="Demographic Parameters",
            category="Phone Number",
            dq_steward=current_user,
            database_name="BBLEDL",
            table_names="tb_income_details",
            param_fields="zmph",
            field_desc="Mobile Phone",
            month=report_month,
            total_records=self.total_records_value,
            missing_records=self.missing_records_phone_number,
            total_qualified=self.total_records_value - self.missing_records_phone_number,
            passed=self.total_records_value - self.missing_records_phone_number,
            failed=self.missing_records_phone_number,
            pass_percentage=None,
            current_score=None,
            expected_score=7,
            threshold=None,
            dimension="Completeness",
            rule_desc="Field cannot be Blank"
        )

        # ✅ Row 9: Address (Validity)
        row09 = self._build_row(
            sub_category="Demographic Parameters",
            category="Address",
            dq_steward=current_user,
            database_name="BBLEDL",
            table_names="tb_contact_details",
            param_fields="pad1,pad2,pad3,pad4, pcity,pzip",
            field_desc="Customer Permanent Address Line(1,2,3,4), city, pin",
            month=report_month,
            total_records=self.total_records_value,
            missing_records=self.missing_record_address,
            total_qualified=self.total_records_value - self.missing_record_address,
            passed=self.total_records_value -self.records_failed_address_validity,
            failed=self.records_failed_address_validity,
            pass_percentage=None,
            current_score=None,
            expected_score=8,
            threshold=None,
            dimension="Validity",
            rule_desc=("At least one address of the borrower should meet all the below conditions: "
                       "1.Address line: a) Minimum length of 5 character State code:"
                       "b) Submission as per catalogue value "
                       "2.Pin code: a) Should be 6 numeric in length "
                       "b) Exclude cases of all digits of same number (0 to 9) "
                       "c) Exclude sequence (E.g., 123456, 456789) "
                       "d) Exclude cases where last 3 digits are numeric 'ZERO`")
        )

        # ✅ Row 10: Address (Completeness)
        row10 = self._build_row(
            sub_category="Demographic Parameters",
            category="Address",
            dq_steward=current_user,
            database_name="BBLEDL",
            table_names="tb_contact_details",
            param_fields="pad1,pad2,pad3,pad4, pcity,pzip",
            field_desc="Customer Permanent Address Line(1,2,3,4), city, pin",
            month=report_month,
            total_records=self.total_records_value,
            missing_records=self.missing_record_address,
            total_qualified=self.total_records_value - self.missing_record_address,
            passed=self.total_records_value - self.missing_record_address,
            failed=self.missing_record_address,
            pass_percentage=None,
            current_score=None,
            expected_score=8,
            threshold=None,
            dimension="Completeness",
            rule_desc="Field cannot be Blank"
        )

        # ✅ Row 11: Total Monthly Family Income (Validity)
        row11 = self._build_row(
            sub_category="Demographic Parameters",
            category="Total Monthly Family Income",
            dq_steward=current_user,
            database_name="BBLEDL",
            table_names="tb_address_details",
            param_fields="cust_income (annual_income), cust_inc  (customer_income_code), cust_inc_desc (income_code_description)",
            field_desc="Customer Annual Income",
            month=report_month,
            total_records=self.total_records_value,
            missing_records=self.missing_record_total_income,
            total_qualified=self.total_records_value - self.missing_record_total_income,
            passed=self.total_records_value - self.records_failed_total_income_validity,
            failed=self.records_failed_total_income_validity,
            pass_percentage=None,
            current_score=None,
            expected_score=5,
            threshold=None,
            dimension="Validity",
            rule_desc=("Availability for all accounts opened on or after April 01, 2022,"
                       " Numeric value except zero with a maximum cap of Rs. 25000/= as per Master Direction on Regulatory Framework for Microfinance Loans Directions,"
                       " 2022 dated March 14, 2022 issued by RBI and amended from time to time")
        )

        # ✅ Row 12: Total Monthly Family Income (Completeness)
        row12 = self._build_row(
            sub_category="Demographic Parameters",
            category="Total Monthly Family Income",
            dq_steward=current_user,
            database_name="BBLEDL",
            table_names="tb_address_details",
            param_fields="cust_income (annual_income), cust_inc  (customer_income_code), cust_inc_desc (income_code_description)",
            field_desc="Customer Annual Income",
            month=report_month,
            total_records=self.total_records_value,
            missing_records=self.missing_record_total_income,
            total_qualified=self.total_records_value - self.missing_record_total_income,
            passed=self.total_records_value - self.missing_record_total_income,
            failed=self.missing_record_total_income,
            pass_percentage=None,
            current_score=None,
            expected_score=5,
            threshold=None,
            dimension="Completeness",
            rule_desc="Field cannot be Blank"
        )

        # Row 13: Key Person Name and Relationship (Validity)
        row13 = self._build_row(
            sub_category="Demographic Parameters",
            category="Key Person Name and Relationship",
            dq_steward=current_user,
            database_name="BBLEDL",
            table_names="tb_cic_relcif_mst_edl, tb_cic_cif_snp_edl",
            param_fields="tb_cic_relcif_mst_edl(acn, poi,  role2 )",
            field_desc="Key Person Name and Relationship",
            month=report_month,
            total_records=self.total_records_value,
            missing_records=self.missing_record_key_person,
            total_qualified=self.total_records_value - self.missing_record_key_person,
            passed=(self.total_records_value - self.missing_record_key_person) - self.records_failed_key_person_validity,
            failed=self.records_failed_key_person_validity,
            pass_percentage=None,
            current_score=None,
            expected_score=5,
            threshold=None,
            dimension="Validity",
            rule_desc="At least 1 key person and relationship"
        )

        # Row 14: Key Person Name and Relationship (Completeness)
        row14 = self._build_row(
            sub_category="Demographic Parameters",
            category="Key Person Name and Relationship",
            dq_steward=current_user,
            database_name="BBLEDL",
            table_names="tb_cic_relcif_mst_edl, tb_cic_cif_snp_edl",
            param_fields="tb_cic_relcif_mst_edl(acn, poi,  role2 )",
            field_desc="Key Person Name and Relationship",
            month=report_month,
            total_records=self.total_records_value,
            missing_records=self.missing_record_key_person,
            total_qualified=self.total_records_value - self.missing_record_key_person,
            passed=self.total_records_value - self.missing_record_key_person,
            failed=self.missing_record_key_person,
            pass_percentage=None,
            current_score=None,
            expected_score=5,
            threshold=None,
            dimension="Completeness",
            rule_desc="Field cannot be Blank"
        )

        return [row01, row02, row03, row04, row05, row06, row07,
                row08, row09, row10, row11, row12, row13, row14]

    def create_excel_workbook(self, headers, data_rows):
        """
        Creates an Excel workbook using OpenPyXL.
        Applies header styling (red font, black background, centered text).
        For each data row:
        - The "Sno" (column A) is auto-incremented.
        - The "Pass %" (column O) is set as a formula that divides the "Records passed"
            by "Total Records Qualified(Active Customers)". Its cell format is set to a percentage.
        The workbook is saved in the instance's xlsx_path.
        """
        # Compose file name dynamically using report file name and current datetime.
        # file_name = RPT_FILE_NAME + '_' + self.current_datetime + '.xlsx'
        file_name = RPT_FILE_NAME + '_' + POST_TXT + '.xlsx'
        wb = Workbook()
        ws = wb.active

        # Define header styles.
        header_font = Font(color="FF0000")  # Red font.
        header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")  # Black background.
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write header row.
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Identify column indexes based on header names.
        sno_index = headers.index("Sno") + 1                            # This will be column A.
        pass_percent_index = headers.index("Pass %") + 1                # Expected to be column O.
        # "Records passed" is column M and "Total Records Qualified(Active Customers)" is column L.
        # Our formula will look like "=M{row_num}/L{row_num}".

        threshold_index = headers.index("Threshold") + 1                # This will be column R.
        current_score_index = headers.index("Current Score") + 1
        passfail_index = headers.index("Pass/Fail") + 1

        # Write data rows.
        for row_idx, row in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                # For "Sno", generate auto-incremented value.
                if col_idx == sno_index:
                    ws.cell(row=row_idx, column=col_idx, value=row_idx - 1)
                # For "Pass %", set the formula result and number format.
                elif col_idx == pass_percent_index:
                    cell = ws.cell(row=row_idx, column=col_idx, value=f"=M{row_idx}/L{row_idx}")
                    cell.number_format = '0.00%'  # ✨✨✨ Format as percentage.

                    """ conditional_formatting pass_percent"""
                    # ✨✨✨
                    col_letter = get_column_letter(pass_percent_index)
                    cell_address = f"{col_letter}{row_idx}"

                    # Rule for - Green background
                    green_fill = PatternFill(bgColor="03FCB6")  # Light green(AAFF00)
                    dxf_green = DifferentialStyle(fill=green_fill)
                    rule_green = Rule(type="expression", dxf=dxf_green)
                    rule_green.formula = [f"={cell_address}>={THRESHOLD}"]
                    ws.conditional_formatting.add(cell_address, rule_green) 

                    # Rule for - Amber background
                    amber_fill = PatternFill(bgColor="F2C5A5")  # Amber/Mustard(FFC125)
                    dxf_amber = DifferentialStyle(fill=amber_fill)
                    rule_amber = Rule(type="expression", dxf=dxf_amber)
                    rule_amber.formula = [f"={cell_address}<{THRESHOLD}"]
                    ws.conditional_formatting.add(cell_address, rule_amber)
                    # ✨✨✨


                elif col_idx == threshold_index:
                    cell = ws.cell(row=row_idx, column=col_idx, value=THRESHOLD)
                    cell.number_format = '0.00%'  # ✨✨✨ Format as percentage.
                elif col_idx == current_score_index:
                    cell = ws.cell(row=row_idx, column=col_idx, value=f"=O{row_idx}*Q{row_idx}")
                elif col_idx == passfail_index:
                    def eval_pass_fail():
                        """
                        eg. =IF(R2<=O2,"Pass",IF(TRUE,"Fail"))
                        """
                        formula = f'''=IF(R{row_idx}<=O{row_idx},"Pass",IF(TRUE,"Fail"))'''
                        # print('formula pass_fail', formula)
                        return formula
                    cell = ws.cell(row=row_idx, column=col_idx, value=eval_pass_fail())

                    """ conditional_formatting Pass/Fail"""
                    # ✨✨✨ 
                    col_letter = get_column_letter(passfail_index)
                    cell_address = f"{col_letter}{row_idx}"

                    # Rule for "Pass" - Green background
                    green_fill = PatternFill(bgColor="03FCB6")  # Light green(AAFF00)
                    dxf_green = DifferentialStyle(fill=green_fill)
                    rule_green = Rule(type="expression", dxf=dxf_green)
                    rule_green.formula = [f'${col_letter}${row_idx}="Pass"']
                    ws.conditional_formatting.add(cell_address, rule_green) 

                    # Rule for "Fail" - Amber background
                    amber_fill = PatternFill(bgColor="F2C5A5")  # Amber/Mustard(FFC125)
                    dxf_amber = DifferentialStyle(fill=amber_fill)
                    rule_amber = Rule(type="expression", dxf=dxf_amber)
                    rule_amber.formula = [f'${col_letter}${row_idx}="Fail"']
                    ws.conditional_formatting.add(cell_address, rule_amber)
                    # ✨✨✨

                else:
                    ws.cell(row=row_idx, column=col_idx, value=value)

        # Add the new conditional formatting rule for data quality rule description.
        self._add_row_wise_conditional_formatting(ws, headers)

        output_file = os.path.join(self.xlsx_path, file_name)
        wb.save(output_file)
        print("Excel file created at", output_file)

    def execute_query(self, conn, query):
        if conn:
            cursor = conn.cursor()
            cursor.execute(query)
            return cursor.status()

    def generate_report(self):
        """
        Orchestrates the complete report generation process:
         - Retrieves the rows (dummy data combined with query results).
         - Defines the headers for the report.
         - Creates the Excel workbook.
        """
        headers = [
            "Sno",                                              # A (auto-generated)
            "SUB-CATEGORY",                                     # B
            "CATEGORY",                                         # C
            "DQ Steward",                                       # D
            "Data base name",                                   # E
            "Table Name(s)",                                    # F
            "Parameters/Field Name(s)",                         # G
            "Field Description",                                # H
            "Month",                                            # I
            "Total Records",                                    # J
            "Missing Records",                                  # K
            "Total Records Qualified(Active Customers)",        # L
            "Records passed",                                   # M
            "Records Failed",                                   # N
            "Pass %",                                           # O - will be a formula here.
            "Current Score",                                    # P
            "Expected Score",                                   # Q
            "Threshold",                                        # R
            "Data Quality Dimension",                           # S
            "Data Quality Rule description",                    # T
            "Pass/Fail",                                        # U
            "Failed record location"                            # V
        ]

        rows = self.get_rows()
        self.create_excel_workbook(headers, rows)


# report_generator = ReportGenerator(SQL_FILE_PATH, XLSX_PATH, SQL_CONTEXT)
# report_generator.generate_report()

if __name__ == "__main__":
    report_generator = ReportGenerator(SQL_FILE_PATH, XLSX_PATH, SQL_CONTEXT)
    report_generator.generate_report()
